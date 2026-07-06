"""
Auto-sync account rules from 1C Excel export.

The sync is intentionally conservative:
- reads one worksheet with fixed semantic columns (company, bank, account);
- adds only missing accounts into ``accounts.json``;
- links new account codes into existing ``companies.json`` entries.
"""

from __future__ import annotations

import json
import logging
import os
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

log = logging.getLogger(__name__)


_BANK_ALIASES: dict[str, str] = {
    "сбер": "sber",
    "сбербанк": "sber",
    "sber": "sber",
    "втб": "vtb",
    "vtb": "vtb",
    "юкб": "ukb",
    "ургиб": "ukb",
    "ургиб банк": "ukb",
    "ukb": "ukb",
    "банк санкт-петербург": "bspb",
    "бспб": "bspb",
    "bspb": "bspb",
    "райффайзен": "raiffeisen",
    "райффайзенбанк": "raiffeisen",
    "raiffeisen": "raiffeisen",
    "точка": "tochka",
    "tochka": "tochka",
    "совкомбанк": "sovcom",
    "совком": "sovcom",
    "sovcom": "sovcom",
}

_PARSER_BY_BANK: dict[str, str] = {
    "sber": "sber_outgoing_balance",
    "vtb": "vtb_outgoing_balance",
    "ukb": "ukb_holder_outgoing",
    "bspb": "bspb_spisanie_balance",
    "raiffeisen": "raiffeisen_outgoing_balance",
    "tochka": "tochka_outgoing_balance",
    "sovcom": "sovcom_outgoing_balance",
}


@dataclass(frozen=True)
class SyncStats:
    added_accounts: int
    linked_company_accounts: int
    skipped_rows: int
    dry_run: bool = False


def _norm_text(s: str) -> str:
    t = (s or "").casefold().strip()
    t = re.sub(r"[\s\"'«»]+", " ", t)
    return t


def _digits_only(s: str) -> str:
    return "".join(ch for ch in str(s or "") if ch.isdigit())


def _bank_code(raw: str) -> str | None:
    n = _norm_text(raw)
    if not n:
        return None
    exact = _BANK_ALIASES.get(n)
    if exact:
        return exact
    # 1C exports full bank names, e.g. "СЕВЕРО-ЗАПАДНЫЙ БАНК ПАО СБЕРБАНК".
    for alias, code in sorted(_BANK_ALIASES.items(), key=lambda x: -len(x[0])):
        if alias in n:
            return code
    return None


def _account_code(base_company_code: str, bank_code: str, account_number: str, existing: set[str]) -> str:
    suffix = account_number[-4:] if len(account_number) >= 4 else account_number
    seed = f"{base_company_code}_{bank_code}_{suffix}"
    if seed not in existing:
        return seed
    # Collision-safe deterministic fallback.
    i = 2
    while True:
        cand = f"{seed}_{i}"
        if cand not in existing:
            return cand
        i += 1


def _match_all_for_bank(bank_code: str, account_number: str) -> list[str]:
    if bank_code == "sber":
        return [account_number, "СберБизнес"]
    if bank_code == "vtb":
        return [account_number, "Владелец счета:"]
    if bank_code == "ukb":
        return [account_number, "holder:"]
    return [account_number]


def _extract_account_digits(rule: dict[str, Any]) -> str:
    return _digits_only(" ".join(str(x) for x in (rule.get("match_all") or [])))


def _ensure_match_all_has_account_number(rule: dict[str, Any], account_number: str) -> None:
    # Tochka PDFs usually expose the account with spaces, not as a contiguous digit string.
    if str(rule.get("bank") or "") == "tochka":
        return
    account_number = _digits_only(account_number)
    if not account_number:
        return
    existing = list(rule.get("match_all") or [])
    if any(account_number in _digits_only(str(x)) for x in existing):
        return
    rule["match_all"] = [account_number, *existing]


def _read_json(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        log.warning("Could not read %s: %s", path, e)
        return []
    return data if isinstance(data, list) else []


def _write_json(path: Path, data: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=4) + "\n",
        encoding="utf-8",
    )


_HEADER_SCAN_MAX_ROWS = 50


def _worksheet_header_index(
    header: list[str], *, strict: bool = True
) -> dict[str, int] | None:
    aliases: dict[str, tuple[str, ...]] = {
        "company": (
            "организация",
            "компания",
            "наименование",
            "контрагент",
            "владелец",
        ),
        "bank": ("банк", "bank"),
        "account": (
            "расчетный счет",
            "расчётный счет",
            "номер счета",
            "номер счёта",
            "счет",
            "счёт",
            "account",
        ),
    }
    out: dict[str, int] = {}
    normalized = [_norm_text(str(c)) for c in header]
    for key, keys in aliases.items():
        for i, val in enumerate(normalized):
            if not val:
                continue
            if any(k in val for k in keys):
                out[key] = i
                break
    missing = [k for k in ("company", "bank", "account") if k not in out]
    if missing:
        if strict:
            raise ValueError(f"Missing required Excel columns: {missing}")
        return None
    return out


def _find_header_row_index(rows: list[tuple[Any, ...]]) -> int | None:
    for i in range(min(_HEADER_SCAN_MAX_ROWS, len(rows))):
        header = [str(c or "") for c in rows[i]]
        if _worksheet_header_index(header, strict=False):
            return i
    return None


def _iter_excel_rows(xlsx_path: Path, sheet_name: str | None = None) -> list[tuple[str, str, str]]:
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    header_row_idx = _find_header_row_index(rows)
    if header_row_idx is None:
        log.warning(
            "Excel %s: no header row with company/bank/account columns in first %d rows (skip sync)",
            xlsx_path,
            _HEADER_SCAN_MAX_ROWS,
        )
        return []
    header = [str(c or "") for c in rows[header_row_idx]]
    idx = _worksheet_header_index(header)
    assert idx is not None
    log.debug("Excel header at row %d: %s", header_row_idx + 1, header)
    out: list[tuple[str, str, str]] = []
    for row in rows[header_row_idx + 1 :]:
        if row is None:
            continue
        company = str(row[idx["company"]] or "").strip()
        bank = str(row[idx["bank"]] or "").strip()
        account = _digits_only(str(row[idx["account"]] or ""))
        if not (company and bank and account):
            continue
        out.append((company, bank, account))
    return out


def sync_accounts_from_excel(
    *,
    excel_path: Path,
    accounts_json_path: Path,
    companies_json_path: Path,
    sheet_name: str | None = None,
    as_of: date | None = None,
    dry_run: bool = False,
) -> SyncStats:
    """
    Merge missing accounts from 1C Excel export into config JSON files.

    Returns counts to help main() decide whether to re-parse unmatched PDFs.
    """
    if not excel_path.is_file():
        log.info("Excel source not found: %s (skip account sync)", excel_path)
        return SyncStats(0, 0, 0, dry_run=dry_run)

    accounts = _read_json(accounts_json_path)
    companies = _read_json(companies_json_path)
    if not companies:
        log.warning("companies.json is empty/invalid; skip Excel sync")
        return SyncStats(0, 0, 0, dry_run=dry_run)

    rows = _iter_excel_rows(excel_path, sheet_name=sheet_name)
    if not rows:
        return SyncStats(0, 0, 0, dry_run=dry_run)

    comp_by_norm_name = {_norm_text(str(c.get("display_name") or "")): c for c in companies}
    existing_codes = {str(a.get("account_code") or "") for a in accounts}
    existing_triplets = {
        (
            str(a.get("company_code") or ""),
            str(a.get("bank") or ""),
            _extract_account_digits(a),
        )
        for a in accounts
    }
    business_date = (as_of or date.today()).strftime("%d.%m.%Y")

    added_accounts = 0
    linked_company_accounts = 0
    skipped_rows = 0
    updated_existing_rules = 0

    dry_run_candidates: list[str] = []
    for company_name, bank_raw, account_number in rows:
        company_obj = comp_by_norm_name.get(_norm_text(company_name))
        if not company_obj:
            skipped_rows += 1
            continue
        bank_code = _bank_code(bank_raw)
        parser = _PARSER_BY_BANK.get(bank_code or "")
        if not bank_code or not parser:
            skipped_rows += 1
            continue
        company_code = str(company_obj.get("company_code") or "")
        if not company_code:
            skipped_rows += 1
            continue

        triplet = (company_code, bank_code, account_number)
        if triplet in existing_triplets:
            continue
        # Legacy manual rules often have (company, bank) but no account digits in match_all.
        # In this case enrich the existing rule instead of creating a duplicate account_code.
        legacy_candidates = [
            a
            for a in accounts
            if str(a.get("company_code") or "") == company_code
            and str(a.get("bank") or "") == bank_code
            and not _extract_account_digits(a)
        ]
        if legacy_candidates:
            before = list(legacy_candidates[0].get("match_all") or [])
            _ensure_match_all_has_account_number(legacy_candidates[0], account_number)
            if list(legacy_candidates[0].get("match_all") or []) != before:
                updated_existing_rules += 1
            existing_triplets.add(triplet)
            continue

        account_code = _account_code(company_code, bank_code, account_number, existing_codes)
        display_name = f'{company_obj.get("display_name", company_code)} {bank_code.upper()}'
        new_account = {
            "account_code": account_code,
            "company_code": company_code,
            "display_name": display_name,
            "bank": bank_code,
            "parser": parser,
            "match_all": _match_all_for_bank(bank_code, account_number),
            "business_date": business_date,
            "active": True,
        }
        accounts.append(new_account)
        existing_codes.add(account_code)
        existing_triplets.add(triplet)
        added_accounts += 1

        acc_list = company_obj.setdefault("accounts", [])
        if account_code not in acc_list:
            acc_list.append(account_code)
            linked_company_accounts += 1
        dry_run_candidates.append(
            f"{account_code} ({company_code}, {bank_code}, ...{account_number[-4:]})"
        )

    if (added_accounts or updated_existing_rules) and not dry_run:
        _write_json(accounts_json_path, accounts)
        if added_accounts:
            _write_json(companies_json_path, companies)
        log.info(
            "Excel sync: added %d account(s), updated %d legacy rule(s), linked %d company account(s), skipped %d row(s)",
            added_accounts,
            updated_existing_rules,
            linked_company_accounts,
            skipped_rows,
        )
    elif (added_accounts or updated_existing_rules) and dry_run:
        log.info(
            "Excel sync DRY-RUN: would add %d account(s), update %d legacy rule(s), link %d company account(s), skipped %d row(s)",
            added_accounts,
            updated_existing_rules,
            linked_company_accounts,
            skipped_rows,
        )
        for cand in dry_run_candidates[:20]:
            log.info("DRY-RUN candidate: %s", cand)
        if len(dry_run_candidates) > 20:
            log.info("DRY-RUN: ... and %d more", len(dry_run_candidates) - 20)

    return SyncStats(added_accounts, linked_company_accounts, skipped_rows, dry_run=dry_run)


def sync_accounts_from_default_excel(
    *,
    accounts_json_path: Path,
    companies_json_path: Path,
) -> SyncStats:
    """
    Convenience wrapper driven by env vars.
    """
    excel_path = Path(
        os.environ.get(
            "VIPISKI_ACCOUNTS_XLSX",
            str(Path(__file__).resolve().parent.parent / "Расчетные счета.xlsx"),
        )
    ).expanduser()
    sheet_name = os.environ.get("VIPISKI_ACCOUNTS_XLSX_SHEET", "").strip() or None
    dry_run = os.environ.get("VIPISKI_ACCOUNTS_SYNC_DRY_RUN", "false").lower() in (
        "1",
        "true",
        "yes",
    )
    return sync_accounts_from_excel(
        excel_path=excel_path,
        accounts_json_path=accounts_json_path,
        companies_json_path=companies_json_path,
        sheet_name=sheet_name,
        dry_run=dry_run,
    )
